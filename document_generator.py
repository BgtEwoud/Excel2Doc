# -*- coding: utf-8 -*-
"""
Created on Thu Dec 28 14:34:14 2023

@author: EwoudBogaert
"""

import os
import glob
from openpyxl import load_workbook
from docxtpl import DocxTemplate
import tkinter as tk
from tkinter import ttk, filedialog
from functools import partial
import pandas as pd

def create_dictionary_from_excel(excel_file_path):
    """Create a dictionary from the first column of the Excel file."""
    df = pd.read_excel(excel_file_path, header=None)
    return {value: None for value in df.iloc[:, 0].tolist()}

def generate_report(docx_file, excel_file, output_folder):
    """Generate reports based on Excel data and a Word template."""
    try:
        to_fill_in = create_dictionary_from_excel(excel_file)
        workbook = load_workbook(excel_file)
        template = DocxTemplate(docx_file)
        worksheet = workbook.active

        for col in range(2, worksheet.max_column + 1):
            for row, key in enumerate(to_fill_in, start=1):
                cell_value = worksheet.cell(row=row, column=col).value
                to_fill_in[key] = cell_value

            template.render(to_fill_in)

            filename = f"{to_fill_in.get('NAAM_VENNOOTSCHAP', 'Unknown')}_{os.path.basename(docx_file).replace('Template', '')}.docx"
            filled_path = os.path.join(output_folder, filename)
            template.save(filled_path)
            print(f"Generated report for: {to_fill_in.get('NAAM_VENNOOTSCHAP', 'Unknown')}")
    except Exception as e:
        print(f"Error processing {docx_file}: {e}")

def select_folder(folder_var, title):
    """Open a folder dialog and update the corresponding variable."""
    folder = filedialog.askdirectory(title=title)
    folder_var.set(folder)

def select_file(file_var, filetypes):
    """Open a file dialog and update the corresponding variable."""
    file = filedialog.askopenfilename(filetypes=filetypes)
    file_var.set(file)

def generate_reports(input_folder_var, excel_file_var, output_folder_var):
    """Generate reports for all DOCX files in the input folder."""
    input_folder = input_folder_var.get()
    excel_file = excel_file_var.get()
    output_folder = output_folder_var.get()
    
    docx_files = glob.glob(os.path.join(input_folder, '*.docx'))
    if not docx_files:
        print("No DOCX files found in the input folder.")
        return

    for docx_file in docx_files:
        print(f"Processing {docx_file}")
        generate_report(docx_file, excel_file, output_folder)

    success_label.config(text="Reports generated successfully!")

def create_gui():
    """Create the GUI for user interaction."""
    global success_label
    root = tk.Tk()
    root.title("Document Generator")

    style = ttk.Style()
    style.configure('TButton', padding=10, relief="flat")

    container = ttk.Frame(root, padding=20)
    container.grid(column=0, row=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    input_folder_var = tk.StringVar()
    excel_file_var = tk.StringVar()
    output_folder_var = tk.StringVar()

    # Input folder
    ttk.Label(container, text="Input Folder:").grid(column=0, row=0, sticky=tk.W)
    ttk.Entry(container, textvariable=input_folder_var).grid(column=0, row=1, columnspan=2, sticky=(tk.W, tk.E))
    ttk.Button(container, text="Browse", command=partial(select_folder, input_folder_var, "Select Input Folder")).grid(column=2, row=1, sticky=tk.W)

    # Excel file
    ttk.Label(container, text="Excel File:").grid(column=0, row=2, sticky=tk.W)
    ttk.Entry(container, textvariable=excel_file_var).grid(column=0, row=3, columnspan=2, sticky=(tk.W, tk.E))
    ttk.Button(container, text="Browse", command=partial(select_file, excel_file_var, [("Excel files", "*.xlsx")])).grid(column=2, row=3, sticky=tk.W)

    # Output folder
    ttk.Label(container, text="Output Folder:").grid(column=0, row=4, sticky=tk.W)
    ttk.Entry(container, textvariable=output_folder_var).grid(column=0, row=5, columnspan=2, sticky=(tk.W, tk.E))
    ttk.Button(container, text="Browse", command=partial(select_folder, output_folder_var, "Select Output Folder")).grid(column=2, row=5, sticky=tk.W)

    # Generate button
    ttk.Button(container, text="Generate Reports", command=partial(generate_reports, input_folder_var, excel_file_var, output_folder_var)).grid(column=0, row=6, columnspan=3, pady=10)

    # Success label
    success_label = ttk.Label(container, text="")
    success_label.grid(column=0, row=7, columnspan=3)

    for child in container.winfo_children():
        child.grid_configure(padx=5, pady=5)

    root.mainloop()

if __name__ == '__main__':
    create_gui()
